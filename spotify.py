# for playlist hot hits 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from bs4 import BeautifulSoup
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# Spotify playlist URL (Replace with your actual playlist URL)
playlist_url = "https://open.spotify.com/playlist/37i9dQZF1DX0XUfTFmNBRM"

# Set up Selenium WebDriver
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)

# Open Spotify playlist
driver.get(playlist_url)

# Wait for the main playlist container to load
wait = WebDriverWait(driver, 10)
wait.until(EC.presence_of_element_located((By.CLASS_NAME, "rezqw3Q4OEPB1m4rmwfw")))

# Scroll using keyboard (simulating real user behavior)
actions = ActionChains(driver)

previous_song_count = 0
target_song_count = 50 

while True:
    actions.send_keys(Keys.PAGE_DOWN).perform()  # Scroll down
    time.sleep(2)  # Wait for content to load

    # Count number of songs loaded
    songs = driver.find_elements(By.CSS_SELECTOR, "div[data-testid='tracklist-row']")
    current_song_count = len(songs)

    print(f"Loaded songs: {current_song_count}")

    # Stop immediately when 50 songs are loaded
    if current_song_count >= target_song_count:
        print("✅ 50 songs loaded. Stopping...")
        break

    # Stop if no new songs are loading
    if current_song_count == previous_song_count:
        print("⚠️ No more songs loading. Stopping...")
        break
    
    previous_song_count = current_song_count

# Extract only the first 50 songs
filtered_songs_html = "".join([song.get_attribute("outerHTML") for song in songs[:50]])

# Save only relevant songs
with open("spotify_1_songs.html", "w", encoding="utf-8") as file:
    file.write(filtered_songs_html)

print("✅ 50 songs saved successfully!")

# Close browser
driver.quit()

with open("spotify_1_songs.html", "r", encoding="utf-8") as f:
    full_data = f.read()
    
soup = BeautifulSoup(full_data, "html.parser")

add_url = "https://open.spotify.com/"

data = []

full_html = soup.find_all("div", {"role": "presentation"})
print(f"Total songs found: {len(full_html)}")

count = 1  # Start from song 1

for html in full_html:
    
    
    # if count > 50:
    #     break  # Stop after processing 50 songs
    
    song_dict = {}
     
    count += 1
    # Fetch song image
    song_image = html.find_all("img", class_="mMx2LUixlnN_Fu45JpFB IqDKYprOtD_EJR1WClPv Yn2Ei5QZn19gria6LjZj")
    for song in song_image:
        song_dict['Image_link'] = song.get("src")
        print(f'Image-Link {song_dict["Image_link"]}')

    # Fetch song name
    song_name = html.find("div", class_="e-9541-text encore-text-body-medium encore-internal-color-text-base btE2c3IKaOXZ4VNAb8WQ standalone-ellipsis-one-line")
    if song_name:  
        song_dict["Song-Name"] = song_name.text.strip()
        print(f"  Song-Name {song_dict['Song-Name']}")

    # Fetch song link
    song_link = html.find_all("a", class_="btE2c3IKaOXZ4VNAb8WQ")
    for song in song_link:
        link = song.get("href")
        song_dict["Song-Link"] = add_url + link
        print(f'Song-Link {song_dict["Song-Link"]}')

    # Fetch album details
    song_album_li = html.find("span", class_="e-9541-text encore-text-body-small")
    if song_album_li:
        song_album = song_album_li.find_all("a", class_="standalone-ellipsis-one-line")
        for album in song_album:
            song_dict["Album-Name"] = album.text.strip()
            link = album.get("href")
            song_dict["Album-Link"] = add_url + link
            print(f' Album-Name {song_dict["Album-Name"]}')
            print(f' Album-Link {song_dict["Album-Link"]}')

    # Fetch artist details
    song_artist_li = html.find("div", class_="e-9541-text encore-text-body-small")
    if song_artist_li:
        song_artists = song_artist_li.find_all("a", href=True)
        for artist in song_artists:
            song_dict["Artist-Name"] = artist.text.strip()
            song_dict["Artist-Link"] = add_url + artist.get("href")
            print(f'Artist-Name {song_dict["Artist-Name"]}')
            print(f'Artist-Link {song_dict["Artist-Link"]}')

    # Fetch date added
    song_date_added = html.find_all("span", class_="e-9541-text encore-text-body-small encore-internal-color-text-subdued standalone-ellipsis-one-line")
    for date in song_date_added:
        song_dict["Date-Added"] = date.text.strip()
        print(f'Date-added {song_dict["Date-Added"]}')

    # Fetch song duration
    song_duration = html.find_all("div", class_="e-9541-text encore-text-body-small encore-internal-color-text-subdued l5CmSxiQaap8rWOOpEpk")
    for song in song_duration:
        song_dict["Duration"] = song.text.strip()
        print(f'Duration {song_dict["Duration"]}')
        
    # Print separator for clarity
    print(f"\n{'=' * 20} Song {count} {'=' * 20}\n")
    # Append song details to list
    data.append(song_dict)
print(f'{"=" *20} "END" {"=" *20} ')
    
df = pd.DataFrame(data)

file_path = "spotify_playlist.xlsx"
name = "Hot Hits Playlist"

# Save to Excel
with pd.ExcelWriter(file_path , engine= "openpyxl") as f:
    df.to_excel(f , index = False , sheet_name= name)
    
print(f"✅ Data successfully saved in '{file_path}' with sheet name '{name}'!")
        
#  open excel file where the data is saved
os.startfile(file_path)  # Opens the file with Excel or the default app

workbook = load_workbook("spotify_playlist.xlsx")
worksheet = workbook.active

hyperlink_col = ["A" , "C" , "E" , "G"]
for col in hyperlink_col:
    for cell in worksheet[col][1:]:   #not taking the header row
        if cell.value and isinstance(cell.value , str) and cell.value.startswith("https"):
            cell.hyperlink = cell.value
            cell.font = Font( color = "0000FF", underline = "single")
            
            
for col in worksheet.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length , len(str(cell.value)))
        except:
            pass
    worksheet.column_dimensions[col_letter].width = max_length + 2

workbook.save("spotify_playlist.xlsx")
